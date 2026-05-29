"""
Validador independent del planning generat. Llegeix un planning xlsx i
comprova contra les regles definides a la Configuració del fitxer d'entrada.

Ús:
    python3 validator.py <entrada.xlsx> <planning.xlsx>
"""
import sys, datetime, re
from collections import defaultdict
import openpyxl
from planning_generator import load_input, _can_work, fold, Config


def parse_planning(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    SHIFT_ROWS = {'8-16': (4, 9), '16-20': (11, 16), '20-08': (18, 23)}
    assignments = {}
    sheets = [s for s in wb.sheetnames if 'setmana' in s.lower()]
    for sn in sheets:
        ws = wb[sn]
        for col in range(2, 9):
            cell = ws.cell(row=3, column=col).value
            if not isinstance(cell, (datetime.date, datetime.datetime)):
                continue
            d = cell.date() if isinstance(cell, datetime.datetime) else cell
            day, dow = d.day, d.weekday()
            for label, (start, end) in SHIFT_ROWS.items():
                cur_group = None
                for r in range(start, end + 1):
                    val = ws.cell(row=r, column=col).value
                    if not val or not isinstance(val, str): continue
                    s = val.strip()
                    if s.startswith(('(', 'Dissabte', 'Diumenge', 'FESTIU', '[')):
                        continue
                    m = re.match(r'^(Grup\s*1|Grup\s*2|N\s+i\s+B|N\s+I\s+B|NB)\s*[:.]?\s*(.*)$', s, re.IGNORECASE)
                    if m:
                        lbl = m.group(1).upper().replace(' ', '')
                        rest = m.group(2).strip()
                        rest = re.sub(r'\s*\(.*\)\s*$', '', rest).strip()
                        if 'GRUP1' in lbl or 'GRUPO1' in lbl:
                            cur_group = 'G1'
                        elif 'GRUP2' in lbl or 'GRUPO2' in lbl:
                            cur_group = 'G2'
                        else:
                            cur_group = 'NB'
                        if rest:
                            if dow in (5, 6) and label == '8-16':
                                slot_label = 'dia (8-20)'
                            elif label == '20-08':
                                slot_label = 'nit (20-08)'
                            else:
                                slot_label = label
                            assignments[(day, dow, slot_label, cur_group)] = rest
    return assignments


def validate(input_path, planning_path):
    constraints, config, meta = load_input(input_path)
    assignments = parse_planning(planning_path)

    errors = []
    warnings = []

    # Match assignment name to constraint name (loose)
    name_set = set(constraints.keys())
    def find_match(name):
        for k in name_set:
            if fold(k) == fold(name): return k
            if fold(k).startswith(fold(name)): return k
        # last name match
        last = fold(name).split(',')[0].split(' ')[0]
        for k in name_set:
            if last in fold(k):
                return k
        return None

    # 1. Double per day
    by_day = defaultdict(list)
    for slot, n in assignments.items():
        by_day[slot[0]].append(n)
    for day, names in by_day.items():
        dup = [n for n in names if names.count(n) > 1]
        if dup:
            errors.append(f"Dia {day}: doble assignació de {set(dup)}")

    # 2. V/C/G
    for slot, name in assignments.items():
        m = find_match(name)
        if m and not _can_work(m, slot[0], constraints):
            cons = constraints.get(m, {})
            reason = cons.get(slot[0]) or 'G adjacent'
            errors.append(f"Dia {slot[0]} {slot[2]} {slot[3]}: {name} té {reason}")

    # 3. Sun-nit-only outside Sun nit
    for slot, name in assignments.items():
        if name in config.sun_nit_only and not (slot[1] == 6 and 'nit' in slot[2]):
            errors.append(f"Dia {slot[0]} {slot[2]}: {name} (Sun-nit-only) fora del seu torn")

    # 4. Weekend-day-only outside Sat/Sun dia
    for slot, name in assignments.items():
        if name in config.weekend_day_only and not (slot[1] in (5, 6) and 'dia' in slot[2]):
            errors.append(f"Dia {slot[0]} {slot[2]}: {name} (Weekend-day-only) fora del seu torn")

    # 5. Equity (pure rotators only)
    exclude = set(config.fix_only) | set(config.sun_nit_only) | set(config.weekend_day_only) | set(config.fix_and_rota)
    counts = defaultdict(int)
    for n in assignments.values():
        m = find_match(n)
        if m and m not in exclude:
            counts[m] += 1
    if counts:
        mx, mn = max(counts.values()), min(counts.values())
        if mx - mn > 1:
            warnings.append(f"Distribució desigual entre rotadors purs: màx {mx}, mín {mn}")

    return errors, warnings, len(assignments)


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(__doc__); sys.exit(2)
    errs, warns, n = validate(sys.argv[1], sys.argv[2])
    print(f"Assignacions: {n}\n")
    if errs:
        print(f"❌ {len(errs)} errors:")
        for e in errs: print(f"  - {e}")
    if warns:
        print(f"\n⚠️  {len(warns)} avisos:")
        for w in warns: print(f"  - {w}")
    if not errs:
        print("✅ Validació OK"); sys.exit(0)
    sys.exit(1)
