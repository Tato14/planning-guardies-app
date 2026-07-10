"""
Generador de planning de guàrdies — versió config-driven.

Sense noms hard-coded al codi. Tota la informació identificativa
(noms, rols, fixos) prové de la plantilla pujada per l'usuari.

Estructura esperada del fitxer d'entrada (xlsx):
  - Pestanya "Configuració": rols fixos + perfils especials
  - Pestanya "Radiòlegs": llista mestra amb rol per professional
  - Pestanya "Vacances": V/C/G/X per dia/professional + metadades

Vegeu 02_Plantilla_Entrada.xlsx per a un exemple buit.
"""
from __future__ import annotations

import calendar
import datetime
import unicodedata
from collections import defaultdict, deque
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# Data structures
# ============================================================

DOW_MAP = {
    'dilluns': 0, 'dimarts': 1, 'dimecres': 2, 'dijous': 3,
    'divendres': 4, 'dissabte': 5, 'diumenge': 6,
    'lunes': 0, 'martes': 1, 'miércoles': 2, 'jueves': 3,
    'viernes': 4, 'sábado': 5, 'domingo': 6,
}


@dataclass
class Config:
    # Fix slots: (dow 0-6, shift '8-16'|'16-20', role 'N'|'B'|'N i B') -> name
    fix_slots: dict = field(default_factory=dict)
    # First-Wed-of-month special slot (16-20)
    first_wed_radiologist: str = ''
    # Lists per role
    rotators: list = field(default_factory=list)       # everyone in rotation
    fix_only: list = field(default_factory=list)       # excluded from rotation
    fix_and_rota: list = field(default_factory=list)   # both
    sun_nit_only: list = field(default_factory=list)
    weekend_day_only: list = field(default_factory=list)
    nou_incorporats: list = field(default_factory=list)  # cobreixen dimecres tarda (excepte 1r Dc del mes)


@dataclass
class Meta:
    year: int
    month: int
    festius: list = field(default_factory=list)
    start_radiologist: str = ''
    first_wed_day: int = 0

    @property
    def days_in_month(self) -> int:
        return calendar.monthrange(self.year, self.month)[1]


# ============================================================
# Helpers
# ============================================================

def fold(s: str) -> str:
    if not s:
        return ''
    return unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode().lower()


def _parse_month(s) -> tuple[int, int]:
    if isinstance(s, datetime.datetime):
        return s.year, s.month
    s = str(s).strip()
    import re
    months = {
        'gener': 1, 'febrer': 2, 'març': 3, 'abril': 4, 'maig': 5, 'juny': 6,
        'juliol': 7, 'agost': 8, 'setembre': 9, 'octubre': 10, 'novembre': 11, 'desembre': 12,
        'enero': 1, 'febrero': 2, 'marzo': 3, 'mayo': 5, 'junio': 6,
        'julio': 7, 'septiembre': 9, 'noviembre': 11, 'diciembre': 12,
        'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
        'july': 7, 'august': 8, 'september': 9, 'october': 10, 'december': 12,
    }
    sl = s.lower()
    year = month = None
    m = re.search(r'(20\d{2})', s)
    if m:
        year = int(m.group(1))
    for k, v in months.items():
        if k in sl:
            month = v
            break
    if month is None:
        m = re.search(r'\b(0?[1-9]|1[0-2])[/\-](20\d{2})\b', s)
        if m:
            month = int(m.group(1))
            year = int(m.group(2))
    if year is None or month is None:
        raise ValueError(f"No s'ha pogut interpretar el mes: '{s}'")
    return year, month


def _parse_festius(s) -> list:
    if not s:
        return []
    parts = [p.strip() for p in str(s).replace(';', ',').split(',') if p.strip()]
    return [int(p) for p in parts if p.isdigit()]


def _first_weekday(year: int, month: int, target_wday: int) -> int:
    for d in range(1, 8):
        if datetime.date(year, month, d).weekday() == target_wday:
            return d
    return 0


# ============================================================
# Loaders
# ============================================================

def load_input(path: str) -> tuple[dict, Config, Meta]:
    """Read the input xlsx and return (constraints, config, meta)."""
    wb = openpyxl.load_workbook(path, data_only=True)

    # ----- Read Radiòlegs sheet -----
    if 'Radiòlegs' not in wb.sheetnames:
        raise ValueError("Falta la pestanya 'Radiòlegs' al fitxer d'entrada.")
    ws_r = wb['Radiòlegs']
    radiologists = {}  # name -> rol
    for row in range(4, ws_r.max_row + 1):
        name = ws_r.cell(row=row, column=1).value
        rol = ws_r.cell(row=row, column=2).value
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        if not name:
            continue
        if rol:
            rol = str(rol).strip().lower()
        else:
            rol = 'rotador'  # default
        radiologists[name] = rol

    # ----- Read Configuració sheet -----
    if 'Configuració' not in wb.sheetnames:
        raise ValueError("Falta la pestanya 'Configuració' al fitxer d'entrada.")
    ws_c = wb['Configuració']

    config = Config()

    # Walk down rows looking for fix slot rows (have a dow name in col A)
    for row in range(4, ws_c.max_row + 1):
        col_a = ws_c.cell(row=row, column=1).value
        col_b = ws_c.cell(row=row, column=2).value  # franja
        col_c = ws_c.cell(row=row, column=3).value  # rol
        col_d = ws_c.cell(row=row, column=4).value  # name

        if not col_a or not isinstance(col_a, str):
            continue
        a = col_a.strip().lower()
        b = (str(col_b).strip().lower() if col_b else '')
        c = (str(col_c).strip() if col_c else '')
        name = (str(col_d).strip() if col_d else '')

        if a in DOW_MAP and b in ('8-16', '16-20') and c and name:
            config.fix_slots[(DOW_MAP[a], b, c)] = name
        elif 'dimecres-1r' in a or '1r-dimecres' in a or 'primer dimecres' in a:
            if name:
                config.first_wed_radiologist = name
        elif a == 'sun-nit-only' and name:
            config.sun_nit_only.append(name)
        elif a == 'weekend-day-only' and name:
            config.weekend_day_only.append(name)

    # Categorize radiologists by rol
    for name, rol in radiologists.items():
        if rol == 'rotador':
            config.rotators.append(name)
        elif rol == 'fix-només' or rol == 'fix-nomes':
            config.fix_only.append(name)
        elif rol == 'fix-i-rota':
            config.fix_and_rota.append(name)
            config.rotators.append(name)  # rotates too
        elif rol == 'sun-nit-only':
            if name not in config.sun_nit_only:
                config.sun_nit_only.append(name)
            config.rotators.append(name)
        elif rol == 'weekend-day-only':
            if name not in config.weekend_day_only:
                config.weekend_day_only.append(name)
            config.rotators.append(name)
        elif rol == 'nou-incorporat' or rol == 'nou incorporat':
            # New professionals: only cover Wed afternoons (except 1st Wed of month)
            # NOT part of general rotation until they graduate to 'rotador'.
            config.nou_incorporats.append(name)

    # ----- Read Vacances sheet -----
    if 'Vacances' not in wb.sheetnames:
        raise ValueError("Falta la pestanya 'Vacances' al fitxer d'entrada.")
    ws_v = wb['Vacances']

    raw_month = ws_v['B1'].value or ''
    raw_start = ws_v['D1'].value or ''
    raw_festius = ws_v['F1'].value or ''

    year, month = _parse_month(raw_month)
    festius = _parse_festius(raw_festius)
    first_wed = _first_weekday(year, month, 2)

    meta = Meta(year=year, month=month, festius=festius,
                start_radiologist=str(raw_start).strip(),
                first_wed_day=first_wed)

    constraints = {}
    for row in range(4, ws_v.max_row + 1):
        name = ws_v.cell(row=row, column=1).value
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        if not name or name not in radiologists:
            continue
        cons = {}
        for day in range(1, 32):
            val = ws_v.cell(row=row, column=1 + day).value
            if val and isinstance(val, str):
                v = val.strip().upper()
                if v in ('V', 'C', 'G', 'X'):
                    cons[day] = v
        constraints[name] = cons

    return constraints, config, meta


# ============================================================
# Algorithm
# ============================================================

def _can_work(name: str, day: int, constraints: dict) -> bool:
    cons = constraints.get(name, {})
    if cons.get(day) in ('V', 'C', 'G'):
        return False
    if day - 1 >= 1 and cons.get(day - 1) == 'G':
        return False
    if cons.get(day + 1) == 'G':
        return False
    return True


def _build_pool(config: Config, start_name: str) -> list:
    pool = sorted(set(config.rotators), key=fold)
    if start_name:
        sf = fold(start_name)
        for i, n in enumerate(pool):
            if sf in fold(n) or fold(n).startswith(sf):
                return pool[i:] + pool[:i]
    return pool


def _gen_slots(meta: Meta) -> list:
    slots = []
    for day in range(1, meta.days_in_month + 1):
        dt = datetime.date(meta.year, meta.month, day)
        dow = dt.weekday()
        is_fest = day in meta.festius
        if is_fest and dow <= 2:
            slots += [(day, dow, 'dia (8-20)', 'NB'),
                      (day, dow, 'nit (20-08)', 'NB')]
        elif is_fest and dow in (3, 4):
            slots += [(day, dow, 'dia (8-20)', 'G1'),
                      (day, dow, 'dia (8-20)', 'G2'),
                      (day, dow, 'nit (20-08)', 'G1'),
                      (day, dow, 'nit (20-08)', 'G2')]
        elif dow in (0, 1):
            slots.append((day, dow, 'nit (20-08)', 'NB'))
        elif dow == 2:
            if day != meta.first_wed_day:
                slots.append((day, dow, '16-20', 'NB'))
            slots.append((day, dow, 'nit (20-08)', 'NB'))
        elif dow in (3, 4):
            slots += [(day, dow, '16-20', 'G1'),
                      (day, dow, '16-20', 'G2'),
                      (day, dow, 'nit (20-08)', 'G1'),
                      (day, dow, 'nit (20-08)', 'G2')]
        elif dow == 5:
            slots += [(day, dow, 'dia (8-20)', 'G1'),
                      (day, dow, 'dia (8-20)', 'G2'),
                      (day, dow, 'nit (20-08)', 'G1'),
                      (day, dow, 'nit (20-08)', 'G2')]
        elif dow == 6:
            slots += [(day, dow, 'dia (8-20)', 'NB'),
                      (day, dow, 'nit (20-08)', 'NB')]
    return slots


def generate_planning(constraints: dict, config: Config, meta: Meta) -> tuple[dict, list, list]:
    warnings = []
    pool = _build_pool(config, meta.start_radiologist)
    if not pool:
        raise ValueError("Pool de rotació buit. Revisa la configuració.")

    slots = _gen_slots(meta)
    assignments = {}
    queue = deque(pool)
    # Separate rotation queue for nou-incorporats (Wed 16-20 slots only)
    nou_queue = deque(sorted(set(config.nou_incorporats), key=fold))
    assigned_today = defaultdict(set)
    deferred_lush = 0
    deferred_arce = 0

    # Pre-populate assigned_today with fixed-slot occupants
    # so that fix-i-rota professionals don't get a rotation slot on a day they cover a fix.
    for day in range(1, meta.days_in_month + 1):
        dt = datetime.date(meta.year, meta.month, day)
        dow = dt.weekday()
        is_fest = day in meta.festius
        if is_fest:
            continue
        # Regular fix slots
        for (d_, lbl, role), name in config.fix_slots.items():
            if d_ != dow: continue
            if not _can_work(name, day, constraints):
                continue
            assigned_today[day].add(name)
        # First Wednesday special slot
        if dow == 2 and day == meta.first_wed_day and config.first_wed_radiologist:
            if _can_work(config.first_wed_radiologist, day, constraints):
                assigned_today[day].add(config.first_wed_radiologist)

    def is_sun_nit(s): return s[1] == 6 and 'nit' in s[2]
    def is_weekend_dia(s): return s[1] in (5, 6) and 'dia' in s[2]
    def is_wed_afternoon(s): return s[1] == 2 and s[2] == '16-20'

    def is_lush(n): return n in config.sun_nit_only
    def is_arce(n): return n in config.weekend_day_only

    for slot in slots:
        day = slot[0]
        # Try to consume pending Lush at Sun nit
        if is_sun_nit(slot) and deferred_lush > 0:
            for lush in config.sun_nit_only:
                if _can_work(lush, day, constraints) and lush not in assigned_today[day]:
                    assignments[slot] = lush
                    assigned_today[day].add(lush)
                    deferred_lush -= 1
                    break
            else:
                pass
            if slot in assignments:
                continue
        # Try Arce at weekend dia
        if is_weekend_dia(slot) and deferred_arce > 0:
            for arce in config.weekend_day_only:
                if _can_work(arce, day, constraints) and arce not in assigned_today[day]:
                    assignments[slot] = arce
                    assigned_today[day].add(arce)
                    deferred_arce -= 1
                    break
            if slot in assignments:
                continue

        # Wed afternoon (16-20, excluding 1st Wed which is a fix): try nou-incorporats first
        if is_wed_afternoon(slot) and nou_queue:
            chosen_nou = None
            tries_nou = 0
            n_nou = len(nou_queue)
            temp_skip_nou = []
            while nou_queue and tries_nou < n_nou:
                cand_nou = nou_queue.popleft()
                if _can_work(cand_nou, day, constraints) and cand_nou not in assigned_today[day]:
                    chosen_nou = cand_nou
                    break
                else:
                    temp_skip_nou.append(cand_nou)
                tries_nou += 1
            # Restore skipped nou-incorporats at the front (they didn't lose their turn)
            for n_ in reversed(temp_skip_nou):
                nou_queue.appendleft(n_)
            if chosen_nou:
                assignments[slot] = chosen_nou
                assigned_today[day].add(chosen_nou)
                nou_queue.append(chosen_nou)  # move to back of nou queue
                continue
            # If no nou-incorporat available, fall through to normal rotation

        # Normal rotation
        chosen = None
        tries = 0
        nq = len(queue)
        temp_skip = []
        while queue and tries < nq:
            cand = queue.popleft()
            if is_lush(cand):
                deferred_lush += 1
                queue.append(cand)
                tries += 1
                continue
            if is_arce(cand):
                deferred_arce += 1
                queue.append(cand)
                tries += 1
                continue
            if _can_work(cand, day, constraints) and cand not in assigned_today[day]:
                chosen = cand
                break
            else:
                temp_skip.append(cand)
            tries += 1
        for n in reversed(temp_skip):
            queue.appendleft(n)
        if chosen is None:
            warnings.append(f"NO trobat per al torn {slot}")
            continue
        assignments[slot] = chosen
        assigned_today[day].add(chosen)
        queue.append(chosen)

    if deferred_lush > 0:
        warnings.append(f"{deferred_lush} torn(s) de Sun-nit-only pendents")
    if deferred_arce > 0:
        warnings.append(f"{deferred_arce} torn(s) de Weekend-day-only pendents")

    return assignments, list(queue), warnings


def last_in_rotation(queue_final: list) -> str:
    return queue_final[-1] if queue_final else ''


# ============================================================
# Writer
# ============================================================

def write_planning(assignments: dict, config: Config, meta: Meta, constraints: dict,
                   template_path: str, output_path: str):
    wb = openpyxl.load_workbook(template_path)
    sheets = [s for s in wb.sheetnames if 'setmana' in s.lower()]

    first_day = datetime.date(meta.year, meta.month, 1)
    week_start = first_day - datetime.timedelta(days=first_day.weekday())

    bold = Font(bold=True, size=10)
    small = Font(size=10)
    fest_f = Font(bold=True, size=10, color='C0392B')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    fill_fix = PatternFill('solid', fgColor='E8F1FB')
    fill_rot = PatternFill('solid', fgColor='F4FBF0')
    fill_fest = PatternFill('solid', fgColor='FFF3CD')
    fill_sub = PatternFill('solid', fgColor='FBEFD8')

    SHIFT_ROWS = {'8-16': (4, 9), '16-20': (11, 16), '20-08': (18, 23)}

    def fixed_for(day, dow, label):
        if day in meta.festius:
            return None
        if label == '16-20' and dow == 2 and day == meta.first_wed_day:
            if config.first_wed_radiologist:
                return [('N i B', f"{config.first_wed_radiologist} (fix 1r dimecres)")]
        result = []
        for role in ('N', 'B', 'N i B'):
            name = config.fix_slots.get((dow, label, role))
            if name:
                # Apply substitution if person can't work
                if not _can_work(name, day, constraints):
                    # Find substitute among fix_and_rota or first available rotator
                    cons = constraints.get(name, {})
                    reason = cons.get(day) or 'G adjacent'
                    name = f"[SUBSTITUIR — {name} té {reason}]"
                result.append((role, name))
        return result if result else None

    def put(ws, col, label, lines, fill):
        start, end = SHIFT_ROWS[label]
        r = start
        for line in lines:
            if r > end:
                break
            c = ws.cell(row=r, column=col)
            c.value = line
            c.alignment = left
            c.font = bold if any(line.startswith(p) for p in ('N:', 'B:', 'N i B', 'Grup', 'Festiu', '(', 'Dissabte', 'Diumenge', 'FESTIU', '[')) else small
            c.fill = fill
            r += 1

    for day in range(1, meta.days_in_month + 1):
        d = datetime.date(meta.year, meta.month, day)
        dow = d.weekday()
        days_offset = (d - week_start).days
        week_idx = days_offset // 7
        if week_idx >= len(sheets):
            continue
        ws = wb[sheets[week_idx]]
        col = 2 + dow
        is_fest = day in meta.festius

        ws.cell(row=3, column=col).value = d
        ws.cell(row=3, column=col).number_format = 'dd/mm (ddd)'
        ws.cell(row=3, column=col).alignment = center
        if is_fest:
            ws.cell(row=3, column=col).font = fest_f

        # 8-16
        if is_fest and dow <= 2:
            nb = assignments.get((day, dow, 'dia (8-20)', 'NB'))
            put(ws, col, '8-16', ['FESTIU — Torn dia 8-20:', f'N i B: {nb}' if nb else '—'], fill_fest)
        elif is_fest and dow in (3, 4):
            g1 = assignments.get((day, dow, 'dia (8-20)', 'G1'))
            g2 = assignments.get((day, dow, 'dia (8-20)', 'G2'))
            put(ws, col, '8-16', ['FESTIU — Torn dia 8-20:', f'Grup 1: {g1}', f'Grup 2: {g2}'], fill_fest)
        elif dow == 5:
            g1 = assignments.get((day, dow, 'dia (8-20)', 'G1'))
            g2 = assignments.get((day, dow, 'dia (8-20)', 'G2'))
            put(ws, col, '8-16', ['Dissabte — Torn dia 8-20:', f'Grup 1: {g1}', f'Grup 2: {g2}'], fill_rot)
        elif dow == 6:
            nb = assignments.get((day, dow, 'dia (8-20)', 'NB'))
            put(ws, col, '8-16', ['Diumenge — Torn dia 8-20:', f'N i B: {nb}'], fill_rot)
        else:
            fx = fixed_for(day, dow, '8-16')
            if fx:
                put(ws, col, '8-16', [f'{r}: {n}' for r, n in fx], fill_fix)

        # 16-20
        if is_fest:
            put(ws, col, '16-20', ['(inclòs torn dia)'], fill_fest)
        elif dow in (5, 6):
            put(ws, col, '16-20', ['(inclòs torn dia)'], fill_rot)
        else:
            fx = fixed_for(day, dow, '16-20')
            if fx:
                has_sub = any('SUBSTITUIR' in n for _, n in fx)
                fill = fill_sub if has_sub else (fill_fest if (dow == 2 and day == meta.first_wed_day) else fill_fix)
                put(ws, col, '16-20', [f'{r}: {n}' for r, n in fx], fill)
            elif dow == 2:
                nb = assignments.get((day, dow, '16-20', 'NB'))
                put(ws, col, '16-20', [f'N i B: {nb}'], fill_rot)
            elif dow in (3, 4):
                g1 = assignments.get((day, dow, '16-20', 'G1'))
                g2 = assignments.get((day, dow, '16-20', 'G2'))
                put(ws, col, '16-20', [f'Grup 1: {g1}', f'Grup 2: {g2}'], fill_rot)

        # 20-08
        if is_fest and dow <= 2:
            nb = assignments.get((day, dow, 'nit (20-08)', 'NB'))
            put(ws, col, '20-08', ['FESTIU — Torn nit 20-08:', f'N i B: {nb}'], fill_fest)
        elif is_fest and dow in (3, 4):
            g1 = assignments.get((day, dow, 'nit (20-08)', 'G1'))
            g2 = assignments.get((day, dow, 'nit (20-08)', 'G2'))
            put(ws, col, '20-08', ['FESTIU — Torn nit 20-08:', f'Grup 1: {g1}', f'Grup 2: {g2}'], fill_fest)
        elif dow in (0, 1, 2):
            nb = assignments.get((day, dow, 'nit (20-08)', 'NB'))
            put(ws, col, '20-08', [f'N i B: {nb}'], fill_rot)
        elif dow in (3, 4, 5):
            g1 = assignments.get((day, dow, 'nit (20-08)', 'G1'))
            g2 = assignments.get((day, dow, 'nit (20-08)', 'G2'))
            put(ws, col, '20-08', [f'Grup 1: {g1}', f'Grup 2: {g2}'], fill_rot)
        elif dow == 6:
            nb = assignments.get((day, dow, 'nit (20-08)', 'NB'))
            put(ws, col, '20-08', [f'N i B: {nb}'], fill_rot)

    for sn in sheets:
        ws = wb[sn]
        ws.column_dimensions['A'].width = 14
        for c in range(2, 9):
            ws.column_dimensions[get_column_letter(c)].width = 32

    wb.save(output_path)
    return output_path


def validate_planning(assignments: dict, config: Config, constraints: dict, meta: Meta) -> list:
    errors = []
    by_day = defaultdict(list)
    for slot, name in assignments.items():
        by_day[slot[0]].append(name)
    for day, names in by_day.items():
        if len(names) != len(set(names)):
            errors.append(f"Dia {day}: doble assignació")
    for slot, name in assignments.items():
        if not _can_work(name, slot[0], constraints):
            errors.append(f"Dia {slot[0]} {slot[2]} {slot[3]}: {name} en conflicte")
    for slot, name in assignments.items():
        if name in config.sun_nit_only and not (slot[1] == 6 and 'nit' in slot[2]):
            errors.append(f"Dia {slot[0]} {slot[2]}: {name} (Sun-nit-only) fora del seu torn")
    for slot, name in assignments.items():
        if name in config.weekend_day_only and not (slot[1] in (5, 6) and 'dia' in slot[2]):
            errors.append(f"Dia {slot[0]} {slot[2]}: {name} (Weekend-day-only) fora del seu torn")
    # Nou-incorporats han d'aparèixer només a dimecres 16-20 (excepte 1r Dc del mes)
    for slot, name in assignments.items():
        if name in config.nou_incorporats:
            if not (slot[1] == 2 and slot[2] == '16-20'):
                errors.append(f"Dia {slot[0]} {slot[2]}: {name} (Nou-incorporat) fora del seu torn Dc 16-20")
            elif slot[0] == meta.first_wed_day:
                errors.append(f"Dia {slot[0]} {slot[2]}: {name} (Nou-incorporat) NO pot cobrir el 1r dimecres del mes")
    # Equity (excluding fix and special)
    counts = defaultdict(int)
    exclude = (set(config.fix_only) | set(config.sun_nit_only) | set(config.weekend_day_only) |
               set(config.fix_and_rota) | set(config.nou_incorporats))
    for n in assignments.values():
        if n not in exclude:
            counts[n] += 1
    if counts:
        mx, mn = max(counts.values()), min(counts.values())
        if mx - mn > 1:
            errors.append(f"Distribució desigual entre rotadors purs: màx {mx}, mín {mn}")
    return errors


# ============================================================
# CLI
# ============================================================
if __name__ == '__main__':
    import argparse, sys
    p = argparse.ArgumentParser()
    p.add_argument('input', help='Fitxer Excel d\'entrada (Configuració + Radiòlegs + Vacances)')
    p.add_argument('template', help='Plantilla Excel buida del planning')
    p.add_argument('output', help='Camí de sortida del planning omplert')
    args = p.parse_args()

    cons, config, meta = load_input(args.input)
    print(f"Mes: {meta.year}-{meta.month:02d}")
    print(f"Rotadors: {len(config.rotators)}, Fixos només: {len(config.fix_only)}")
    print(f"Sun-nit-only: {config.sun_nit_only}")
    print(f"Weekend-day-only: {config.weekend_day_only}")
    print(f"Nou-incorporats (dimecres 16-20): {config.nou_incorporats}")

    assignments, queue, warn = generate_planning(cons, config, meta)
    write_planning(assignments, config, meta, cons, args.template, args.output)
    errs = validate_planning(assignments, config, cons, meta)

    print(f"\nAssignacions: {len(assignments)}")
    print(f"Últim de la roda: {last_in_rotation(queue)}")
    if warn:
        print("\nAvisos:")
        for w in warn: print(f"  - {w}")
    if errs:
        print("\nErrors:")
        for e in errs: print(f"  - {e}")
        sys.exit(1)
    print("\nValidació OK ✓")
