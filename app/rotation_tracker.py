"""
Detecta l'últim radiòleg de la roda d'un planning anterior i suggereix
qui ha de començar la roda del mes nou.

Ús:
    python3 rotation_tracker.py <planning_mes_anterior.xlsx> [entrada_mes_nou.xlsx]
"""
import sys, datetime, re
from openpyxl import load_workbook
from planning_generator import fold, load_input


def find_last_rotator(planning_path, exclude_set):
    wb = load_workbook(planning_path, data_only=True)
    sheets = [s for s in wb.sheetnames if 'setmana' in s.lower()]
    last_date = None
    last_name = None
    for sn in sheets:
        ws = wb[sn]
        for col in range(2, 9):
            cell = ws.cell(row=3, column=col).value
            if not isinstance(cell, (datetime.date, datetime.datetime)):
                continue
            d = cell.date() if isinstance(cell, datetime.datetime) else cell
            for r in range(4, 24):
                v = ws.cell(row=r, column=col).value
                if not v or not isinstance(v, str): continue
                s = v.strip()
                m = re.match(r'^(Grup\s*[12]|N\s+i\s+B|N\s+I\s+B|NB)\s*[:.]?\s*(.+)$', s, re.IGNORECASE)
                if m:
                    name = m.group(2).strip()
                    name = re.sub(r'\s*\(.*\)\s*$', '', name).strip()
                    if fold(name) in (fold(x) for x in exclude_set):
                        continue
                    if last_date is None or d > last_date:
                        last_date = d
                        last_name = name
    return last_name, last_date


def next_in_alphabet(name, all_names):
    pool = sorted(set(all_names), key=fold)
    nf = fold(name)
    for i, n in enumerate(pool):
        if fold(n) == nf or fold(n).startswith(nf):
            return pool[(i + 1) % len(pool)]
    for n in pool:
        if fold(n) > nf: return n
    return pool[0]


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(2)
    planning = sys.argv[1]

    if len(sys.argv) >= 3:
        _, config, _ = load_input(sys.argv[2])
        exclude = set(config.fix_only) | set(config.sun_nit_only) | set(config.weekend_day_only)
        last_name, last_date = find_last_rotator(planning, exclude)
        if last_name:
            print(f"Últim rotador del mes anterior: {last_name}")
            print(f"  (data: {last_date})")
            pool = list(set(config.rotators))
            nxt = next_in_alphabet(last_name, pool)
            print(f"\nPrimer de la roda del mes nou: {nxt}")
        else:
            print("No s'ha detectat l'últim rotador.")
            sys.exit(1)
    else:
        last_name, last_date = find_last_rotator(planning, set())
        if last_name:
            print(f"Últim del planning: {last_name} ({last_date})")
            print("Per saber el primer del mes següent, passa també el fitxer d'entrada del mes nou.")
        else:
            print("No s'ha detectat l'últim. Revisa el fitxer.")
            sys.exit(1)
